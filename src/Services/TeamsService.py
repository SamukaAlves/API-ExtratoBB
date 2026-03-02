import openpyxl
import requests
import json
from datetime import datetime
from typing import Dict, Optional

class TeamsService:
    def __init__(self, webhook_url: str, id_bot: Optional[int] = None):
        # garantir que o console Python esteja em utf-8 (windows)
        try:
            import sys
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass

        self.webhook_url = webhook_url
        self.id_bot = id_bot
    
    def _celula_e_data(self, val):
        """Retorna (True, str_data) se o valor for data (datetime ou dd/mm/yyyy), senão (False, None)."""
        if val is None:
            return False, None
        if isinstance(val, datetime):
            return True, val.strftime('%d/%m/%Y')
        s = str(val).strip()
        if len(s) == 10 and s[2] == '/' and s[5] == '/':
            try:
                datetime.strptime(s, '%d/%m/%Y')
                return True, s
            except ValueError:
                pass
        return False, None

    def processar_excel_extrato(self, caminho_excel: str) -> Optional[Dict]:
        """Processa o arquivo Excel do extrato.
        Estrutura real: A=Data, B=obs, C=Data balancete, D=Agencia Origem, E=Lote,
        F=Numero Documento, G=Cod.Historico, H=Historico, I=Valor R$, J=Inf.(C/D), K=Detalhamento.
        Saldo atual: linha onde A contém 'Saldo' e valor em B."""
        try:
            wb = openpyxl.load_workbook(caminho_excel, data_only=True)
            ws = wb.active
            
            dados = {
                'agencia': None,
                'conta': None,
                'saldo_anterior': None,
                'lancamentos': [],
                'saldo_atual': None,
                'data_debito_juros': None,
                'data_debito_iof': None
            }
            
            for row_idx in range(1, ws.max_row + 1):
                cell_a = ws.cell(row_idx, 1).value
                cell_b = ws.cell(row_idx, 2).value
                cell_d = ws.cell(row_idx, 4).value   # Agencia Origem
                cell_f = ws.cell(row_idx, 6).value   # Numero Documento
                cell_h = ws.cell(row_idx, 8).value   # Historico
                cell_i = ws.cell(row_idx, 9).value   # Valor R$
                cell_j = ws.cell(row_idx, 10).value  # Inf. C/D
                cell_k = ws.cell(row_idx, 11).value  # Detalhamento
            
                # Agência: linha com "Agencia" em A, número em B
                if cell_a and 'Agencia' in str(cell_a).strip():
                    if cell_b is not None:
                        num = str(cell_b).replace(' ', '').replace('-', '')
                        if num.isdigit() and len(num) >= 4:
                            dados['agencia'] = f"{num[:4]}-{num[4]}" if len(num) > 4 else num
                
                # Conta corrente: no layout BB, linha 2 tem C="Conta corrente" e D=número
                cell_c = ws.cell(row_idx, 3).value
                if cell_c and 'Conta corrente' in str(cell_c).strip():
                    conta_val = ws.cell(row_idx, 4).value
                    if conta_val is not None:
                        num = str(conta_val).replace(' ', '').replace('-', '')
                        if num.isdigit() and len(num) >= 2:
                            dados['conta'] = f"{num[:-1]}-{num[-1]}" if len(num) > 1 else num
                
                # Linhas de dados (A = data)
                eh_data, data_str = self._celula_e_data(cell_a)
                if eh_data and cell_h is not None:
                    historico = str(cell_h).strip()
                    
                    # Saldo Anterior: valor na coluna I, tipo na J
                    if 'Saldo Anterior' in historico:
                        if cell_i is not None:
                            valor_tipo = f"{str(cell_i).strip()} {str(cell_j or '').strip()}".strip()
                            dados['saldo_anterior'] = self._extrair_valor_da_celula(valor_tipo or str(cell_i))
                        continue
                    
                    # Pular linha de S A L D O (o saldo final vem na seção “Saldo” em A/B)
                    if 'S A L D O' in historico or (historico == 'SALDO' and cell_i is None):
                        continue
                    
                    # Lançamento: valor na coluna I, tipo na J; documento na F
                    valor_tipo = f"{str(cell_i).strip()} {str(cell_j or '').strip()}".strip() if cell_i is not None else None
                    valor = self._extrair_valor_da_celula(valor_tipo) if valor_tipo else None
                    documento = str(cell_f).strip() if cell_f else ''
                    detalhamento = str(cell_k)[:80] if cell_k else ''
                    
                    if historico and valor:
                        dados['lancamentos'].append({
                            'data': data_str,
                            'historico': historico,
                            'documento': documento,
                            'valor': valor,
                            'detalhamento': detalhamento
                        })
                
                # Saldo atual: rótulo "Saldo" ou "Saldo Atual" na coluna A, valor na coluna B
                if cell_a is not None:
                    sa = str(cell_a).strip()
                    if sa in ('Saldo', 'Saldo Atual') or 'Saldo Atual' in sa:
                        saldo_b = ws.cell(row_idx, 2).value
                        if saldo_b is not None and str(saldo_b).strip():
                            dados['saldo_atual'] = self._extrair_valor_da_celula(str(saldo_b).strip())
                
                # Data de Débito de Juros: rótulo em A, valor em B
                if cell_a and 'Data de Debito de Juros' in str(cell_a):
                    v = ws.cell(row_idx, 2).value
                    if v:
                        dados['data_debito_juros'] = v.strftime('%d/%m/%Y') if isinstance(v, datetime) else str(v).strip()
                
                # Data de Débito de IOF: rótulo em A, valor em B
                if cell_a and 'Data de Debito de IOF' in str(cell_a):
                    v = ws.cell(row_idx, 2).value
                    if v:
                        dados['data_debito_iof'] = v.strftime('%d/%m/%Y') if isinstance(v, datetime) else str(v).strip()
            
            return dados
            
        except Exception as e:
            print(f"Erro ao processar Excel: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extrair_valor_da_celula(self, cell_value) -> Dict:
        """Extrai valor numérico e tipo (C/D) de uma célula. Aceita '1.234,56 C' ou só '1.234,56'."""
        try:
            cell_str = str(cell_value).strip()
            if not cell_str:
                return {'valor': 0, 'tipo': 'C', 'formatado': '0,00'}
            
            # Identificar tipo: C/D no fim da string; se não tiver, assume C
            up = cell_str.upper()
            if ' C' in up or up.endswith('C'):
                tipo = 'C'
                valor_str = cell_str.replace('C', '').replace('c', '').strip()
            elif ' D' in up or up.endswith('D'):
                tipo = 'D'
                valor_str = cell_str.replace('D', '').replace('d', '').strip()
            else:
                tipo = 'C'
                valor_str = cell_str
            
            # Formato BR: 1.261.073,38 -> remover pontos, trocar vírgula por ponto
            valor_str = valor_str.replace('.', '').replace(',', '.').replace(' ', '')
            valor_float = float(valor_str)
            
            return {
                'valor': valor_float,
                'tipo': tipo,
                'formatado': self._formatar_valor(valor_float)
            }
        except Exception as e:
            print(f"Erro ao extrair valor de '{cell_value}': {e}")
            return {'valor': 0, 'tipo': 'C', 'formatado': '0,00'}
    
    def _formatar_valor(self, valor) -> str:
        """Formata valor para padrão brasileiro"""
        try:
            if isinstance(valor, str):
                valor = valor.replace('.', '').replace(',', '.')
            valor_float = float(valor)
            
            formatted = f"{valor_float:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
            return formatted
        except:
            return "0,00"
    
    def formatar_extrato_txt(self, dados: Dict, carteira: str) -> str:
        """Formata extrato de forma compacta para Teams: alinhado, sem estourar largura, legível no celular."""
        # Largura restrita (~59 chars) para não quebrar no mobile e colunas não se misturarem
        W = 59
        sep = "=" * (W - 1)
        div = "-" * (W - 1)
        t = []
        
        # Cabeçalho
        t.append(sep)
        t.append(f"Extrato conta corrente - {carteira}")
        t.append(datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        t.append(div)
        
        # Dados da conta
        t.append(f"Agência: {dados['agencia'] or 'N/A'}")
        t.append(f"Conta: {dados['conta'] or 'N/A'}")
        t.append(div)
        
        # Cabeçalho da tabela: colunas fixas (Data 10, Hist 22, Doc 10, Valor+C/D 14)
        t.append("Lançamentos")
        t.append(div)
        t.append(f"{'Data':<10} {'Histórico':<22} {'Doc':<10} {'Valor R$':>14}")
        t.append(div)
        
        # Saldo Anterior
        if dados.get("saldo_anterior"):
            sa = dados["saldo_anterior"]
            t.append(f"{'Saldo Anterior':<22} {'':<10} {sa['formatado']:>12} {sa['tipo']}")
        
        # Lançamentos
        for lanc in dados.get("lancamentos", []):
            v = lanc["valor"]
            hist = (lanc.get('historico') or '')[:22]
            doc = (lanc.get('documento') or '')[:10]
            t.append(f"{lanc.get('data',''):<10} {hist:<22} {doc:<10} {v['formatado']:>12} {v['tipo']}")
            if lanc.get('detalhamento') and str(lanc['detalhamento']).strip():
                det = str(lanc['detalhamento'])[:54].strip()
                t.append(f"  + {det}")
        
        t.append(div)
        
        # Resumo
        if dados.get("saldo_atual"):
            s = dados["saldo_atual"]
            t.append(f"Saldo Atual: {s['formatado']} {s['tipo']}")
        if dados.get("data_debito_juros"):
            t.append(f"Dt.Déb.Juros: {dados['data_debito_juros']}")
        if dados.get("data_debito_iof"):
            t.append(f"Dt.Déb.IOF: {dados['data_debito_iof']}")
        t.append(div)
        
        return "\n".join(t)

    def _montar_adaptive_card_extrato(self, dados: Dict, carteira: str) -> list:
        """Monta o body do Adaptive Card: formato igual ao extrato em texto, só os valores coloridos (C=azul, D=vermelho)."""
        W = 59
        sep = "=" * (W - 1)
        div = "-" * (W - 1)
        body = []

        def _line(txt: str):
            """Uma linha em monospace, cor padrão."""
            body.append({"type": "TextBlock", "text": txt, "wrap": True, "fontType": "monospace"})

        def _line_com_valor(prefixo: str, valor_formatado: str, tipo: str, alinhar_valor: bool = True):
            """Uma linha em que só o valor (número + C/D) tem cor: C=azul, D=vermelho."""
            cor = "accent" if tipo == "C" else "attention"
            valor_txt = f"{valor_formatado:>12} {tipo}" if alinhar_valor else f"{valor_formatado} {tipo}"
            body.append({
                "type": "RichTextBlock",
                "inlines": [
                    {"type": "TextRun", "text": prefixo, "fontType": "monospace"},
                    {"type": "TextRun", "text": valor_txt, "fontType": "monospace", "color": cor}
                ]
            })

        # Título (separador com 9 "=" a menos que o restante)
        titulo = f"Extrato Bancário - {datetime.now().strftime('%d/%m/%Y - %H:%M')}"
        _line(titulo)
        _line("=" * (W - 1 - 9))
        _line(f"Extrato conta corrente - {carteira}")
        _line(datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        _line(div)

        _line(f"Agência: {dados['agencia'] or 'N/A'}")
        _line(f"Conta: {dados['conta'] or 'N/A'}")
        _line(div)

        _line("Lançamentos")
        _line(div)
        _line(f"{'Data':<10} {'Histórico':<22} {'Doc':<10} {'Valor R$':>14}")
        _line(div)

        # Saldo Anterior: prefixo fixo, só o valor colorido
        if dados.get("saldo_anterior"):
            sa = dados["saldo_anterior"]
            prefixo_sa = f"{'Saldo Anterior':<22} {'':<10} "
            _line_com_valor(prefixo_sa, sa['formatado'], sa.get('tipo', 'C'))

        # Lançamentos: prefixo = data + hist + doc; só valor colorido
        for lanc in dados.get("lancamentos", []):
            v = lanc["valor"]
            hist = (lanc.get('historico') or '')[:22]
            doc = (lanc.get('documento') or '')[:10]
            prefixo = f"{lanc.get('data',''):<10} {hist:<22} {doc:<10} "
            _line_com_valor(prefixo, v['formatado'], v.get('tipo', 'C'))
            if lanc.get('detalhamento') and str(lanc['detalhamento']).strip():
                det = str(lanc['detalhamento'])[:54].strip()
                _line(f"  + {det}")

        _line(div)

        # Saldo Atual: prefixo "Saldo Atual: ", só valor colorido (sem alinhamento fixo)
        if dados.get("saldo_atual"):
            s = dados["saldo_atual"]
            _line_com_valor("Saldo Atual: ", s['formatado'], s.get('tipo', 'C'), alinhar_valor=False)
        if dados.get("data_debito_juros"):
            _line(f"Dt.Déb.Juros: {dados['data_debito_juros']}")
        if dados.get("data_debito_iof"):
            _line(f"Dt.Déb.IOF: {dados['data_debito_iof']}")
        _line(div)

        return body

    def _safe_print(self, text: str):
        """Print que nunca falha por encoding; substitui caracteres inválidos."""
        try:
            print(text)
        except UnicodeEncodeError:
            # forçar ambiente a utf-8 e substituir
            try:
                import sys
                sys.stdout.reconfigure(encoding='utf-8', errors='replace')
            except Exception:
                pass
            print(text.encode('utf-8', errors='replace').decode('utf-8'))

    def enviar_extrato_teams(self, dados: Dict, carteira: str):
        """Envia extrato para Teams como Adaptive Card (créditos em azul, débitos em vermelho)."""
        if not self.webhook_url:
            self._safe_print("Webhook do Teams não configurado")
            return False

        body = self._montar_adaptive_card_extrato(dados, carteira)
        card = {
            "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
            "type": "AdaptiveCard",
            "version": "1.2",
            "body": body
        }
        mensagem = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "contentUrl": None,
                    "content": card
                }
            ]
        }

        try:
            response = requests.post(
                self.webhook_url,
                headers={"Content-Type": "application/json"},
                data=json.dumps(mensagem),
                timeout=10
            )

            if response.status_code == 200:
                self._safe_print("Extrato enviado para o Teams!")
                # Atualiza última execução se id_bot informado
                try:
                    if self.id_bot:
                        from src.Repository.masterSQLCom import MasterSQLComunication
                        MasterSQLComunication().upload_log(self.id_bot)
                except Exception:
                    pass
                return True
            else:
                self._safe_print(f"Erro: {response.status_code}")
                self._safe_print(f"   Resposta: {response.text}")
                return False
        except Exception as e:
            self._safe_print(f"Erro: {e}")
            return False
    
    def enviar_mensagem_erro(self, erro: str):
        """Envia notificação de erro"""
        if not self.webhook_url:
            return
        
        mensagem = {
            "@type": "MessageCard",
            "@context": "https://schema.org/extensions",
            "themeColor": "FF0000",
            "title": "Erro no Bot de Extratos",
            "text": f"```\n{erro[:500]}\n```"
        }
        
        try:
            requests.post(self.webhook_url, headers={"Content-Type": "application/json"}, 
                    data=json.dumps(mensagem), timeout=10)
        except Exception as e:
            self._safe_print(f"Falha ao enviar mensagem de erro: {e}")